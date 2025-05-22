from pathlib import Path
import os
from fastapi import FastAPI, Query
from pydantic import BaseModel
from typing import List
import openpyxl

app = FastAPI()

CARPETA_MODULO = Path(__file__).parent
ARCHIVO_EXCEL = str(CARPETA_MODULO / "registros.xlsx")
AUDITORIA_EXCEL = str(CARPETA_MODULO / "auditoria.xlsx")

# Modelo de datos
class Contador(BaseModel):
    fecha: str
    casino: str
    maquina: str
    in_contador: float
    out_contador: float
    jackpot_contador: float
    billetero_contador: float

# Inicializar archivo Excel si no existe
def inicializar_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha", "Casino", "Maquina", "In", "Out", "Jackpot", "Billetero"])
        wb.save(ARCHIVO_EXCEL)

inicializar_excel()

# Guardar nuevo registro
@app.post("/registrar")
def registrar_contador(contador: Contador):
    for valor in [contador.in_contador, contador.out_contador, contador.jackpot_contador, contador.billetero_contador]:
        if valor < 0:
            return {"error": "Los contadores no pueden tener valores negativos"}

    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    ws.append([
        contador.fecha,
        contador.casino,
        contador.maquina,
        contador.in_contador,
        contador.out_contador,
        contador.jackpot_contador,
        contador.billetero_contador
    ])
    wb.save(ARCHIVO_EXCEL)
    return {"mensaje": "Registro guardado exitosamente"}

# Obtener todos los registros
@app.get("/registros")
def obtener_registros():
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    registros = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        registros.append({
            "fecha": row[0],
            "casino": row[1],
            "maquina": row[2],
            "in": row[3],
            "out": row[4],
            "jackpot": row[5],
            "billetero": row[6]
        })
    if not registros:
        return {"Mensaje": "No hay registros BB"}
    return registros

# Buscar por casino y fecha
@app.get("/buscar/")
def buscar_registros(casino: str = Query(...), fecha: str = Query(...)):
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    resultados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == casino and row[0] == fecha:
            resultados.append({
                "fecha": row[0],
                "casino": row[1],
                "maquina": row[2],
                "in": row[3],
                "out": row[4],
                "jackpot": row[5],
                "billetero": row[6]
            })
    return resultados

# Modificar registro
@app.put("/modificar")
def modificar_contador(contador: Contador):
    for valor in [contador.in_contador, contador.out_contador, contador.jackpot_contador, contador.billetero_contador]:
        if valor < 0:
            return {"error": "Los contadores no pueden tener valores negativos"}

    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    modificado = False

    for row in ws.iter_rows(min_row=2):
        if (row[0].value == contador.fecha and
            row[1].value == contador.casino ):

            anterior = [cell.value for cell in row]
            row[3].value = contador.in_contador
            row[4].value = contador.out_contador
            row[5].value = contador.jackpot_contador
            row[6].value = contador.billetero_contador
            modificado = True

            registrar_auditoria(anterior, contador)
            break

    if not modificado:
        return {"error": "Registro no encontrado"}

    wb.save(ARCHIVO_EXCEL)
    return {"mensaje": "Registro modificado exitosamente"}

# Registrar auditoría
def registrar_auditoria(anterior, nuevo: Contador):
    if not os.path.exists(AUDITORIA_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "casino", "maquina", "in_antes", "out_antes", "jackpot_antes", "billetero_antes",
                   "in_despues", "out_despues", "jackpot_despues", "billetero_despues"])
        wb.save(AUDITORIA_EXCEL)

    wb = openpyxl.load_workbook(AUDITORIA_EXCEL)
    ws = wb.active
    ws.append([
        anterior[0], anterior[1], anterior[2],
        anterior[3], anterior[4], anterior[5], anterior[6],
        nuevo.in_contador, nuevo.out_contador,
        nuevo.jackpot_contador, nuevo.billetero_contador
    ])
    wb.save(AUDITORIA_EXCEL)

# uvicorn Registro_Contadores:app --reload