import os
import uuid
import pandas as pd
from fpdf import FPDF
from fastapi.responses import FileResponse

class ExportadorReportes:

    @staticmethod
    def exportar_a_pdf(data: dict, nombre_archivo: str) -> str:
        
        from fpdf import FPDF
        import os

        os.makedirs("exports", exist_ok=True)
        ruta = f"exports/{nombre_archivo}.pdf"

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", style='B', size=14)
        pdf.cell(200, 10, txt="REPORTE DE UTILIDAD", ln=True, align='C')
        pdf.ln(8)

    # Encabezado general
        pdf.set_font("Arial", size=11)
        fecha_ini = data.get("fecha_inicio", "")
        fecha_fin = data.get("fecha_fin", "")
        casino = data.get("casino", "")
        tipo = data.get("tipo", "reporte").upper()
        pdf.cell(200, 10, txt=f"Tipo de reporte: {tipo}", ln=True)
        pdf.cell(200, 10, txt=f"Casino: {casino}", ln=True)
        pdf.cell(200, 10, txt=f"Rango de fechas: {fecha_ini} a {fecha_fin}", ln=True)
        pdf.ln(5)

    # Detalle por máquina
        detalle = data.get("detalle_maquinas", [])
        if not detalle:
            pdf.cell(200, 10, txt="Sin datos disponibles para mostrar.", ln=True)
        else:
            for maquina in detalle:
                
                pdf.set_font("Arial", style='B', size=12)
                pdf.cell(200, 8, txt=f"Máquina: {maquina.get('maquina', 'Sin datos')}", ln=True)
                pdf.set_font("Arial", size=11)
                pdf.cell(200, 8, txt=f"Denominación: {maquina.get('denominacion', '-')}", ln=True)
                pdf.cell(200, 8, txt=f"IN: {maquina.get('total_in', '-')}", ln=True)
                pdf.cell(200, 8, txt=f"OUT: {maquina.get('total_out', '-')}", ln=True)
                pdf.cell(200, 8, txt=f"JACKPOT: {maquina.get('total_jackpot', '-')}", ln=True)
                pdf.cell(200, 8, txt=f"BILLETERO: {maquina.get('total_billetero', '-')}", ln=True)
                pdf.cell(200, 8, txt=f"UTILIDAD: {maquina.get('utilidad', '-')}", ln=True)
                pdf.ln(5)

    # Totales generales
        pdf.set_font("Arial", style='B', size=12)
        pdf.ln(5)
        pdf.cell(200, 10, txt="TOTALES GENERALES:", ln=True)
        pdf.set_font("Arial", size=11)
        totales = data.get("totales_grupo") or data.get("totales") or {}
        for clave, valor in totales.items():
            pdf.cell(200, 8, txt=f"{clave.upper()}: {valor}", ln=True)

        pdf.output(ruta)
        return ruta
    
    @staticmethod
    def exportar_a_excel(data: dict, nombre_archivo: str) -> str:
        
        os.makedirs("exports", exist_ok=True)
        ruta = f"exports/{nombre_archivo}.xlsx"

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Utilidad"

    # Estilos
        bold = Font(bold=True)
        center = Alignment(horizontal="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill("solid", fgColor="DDDDDD")
        fill_total = PatternFill("solid", fgColor="CCE5FF")

    # Encabezado
        ws["A1"] = "REPORTE DE UTILIDAD"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:C1")
        ws["A1"].alignment = center

        fila = 3
        tipo_reporte = data.get("tipo_reporte", "").upper()
        casino = data.get("casino", "")
        fecha_inicio = data.get("fecha_inicio", "")
        fecha_fin = data.get("fecha_fin", "")

        ws[f"A{fila}"] = f"Tipo de reporte:"
        ws[f"B{fila}"] = tipo_reporte
        fila += 1
        ws[f"A{fila}"] = f"Casino:"
        ws[f"B{fila}"] = casino
        fila += 1
        ws[f"A{fila}"] = f"Rango de fechas:"
        ws[f"B{fila}"] = f"{fecha_inicio} a {fecha_fin}"
        fila += 2

        for maquina in data.get("detalle_maquinas", []):
            ws[f"A{fila}"] = f"Máquina:"
            ws[f"B{fila}"] = maquina.get("maquina", "Sin datos")
            fila += 1
            ws[f"A{fila}"] = "Denominación:"
            ws[f"B{fila}"] = maquina.get("denominacion", "")
            fila += 1

            etiquetas = ["IN", "OUT", "JACKPOT", "BILLETERO", "UTILIDAD"]
            claves = ["total_in", "total_out", "total_jackpot", "total_billetero", "utilidad"]

            for et, key in zip(etiquetas, claves):
                ws[f"A{fila}"] = et
                ws[f"B{fila}"] = maquina.get(key, "")
                ws[f"A{fila}"].font = bold
                ws[f"A{fila}"].fill = fill_header
                ws[f"A{fila}"].alignment = center
                ws[f"A{fila}"].border = border
                ws[f"B{fila}"].border = border
                fila += 1

            fila += 1

    # Totales
        ws[f"A{fila}"] = "TOTALES GENERALES:"
        ws[f"A{fila}"].font = Font(bold=True)
        fila += 1

        totales = data.get("totales_grupo") or data.get("totales") or {}
        for clave, valor in totales.items():
            ws[f"A{fila}"] = clave.upper()
            ws[f"B{fila}"] = valor
            ws[f"A{fila}"].fill = fill_total
            ws[f"B{fila}"].fill = fill_total
            ws[f"A{fila}"].font = bold
            ws[f"B{fila}"].font = bold
            ws[f"A{fila}"].border = border
            ws[f"B{fila}"].border = border
            fila += 1

        for col in range(1, 4):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        wb.save(ruta)
        return ruta


    @staticmethod
    def generar_nombre_archivo() -> str:
        return f"reporte_{uuid.uuid4().hex[:8]}"

    @staticmethod
    def devolver_archivo(ruta: str) -> FileResponse:
        return FileResponse(path=ruta, filename=os.path.basename(ruta), media_type="application/octet-stream")

