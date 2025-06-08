from fastapi import APIRouter, Query
from fastapi.responses import FileResponse, JSONResponse
from typing import List, Optional
from datetime import date
import os
import pandas as pd
from fpdf import FPDF
import uuid
from app.modules.cuadre_casino.cuadre_casino import cuadre_casino, CuadreCasinoRequest
from app.modules.cuadre_maquina.cuadre_maquina import calcular_balance, BalanceRequest
from app.modules.gestion_lugares.lugares import cargar_lugares
import json
from pathlib import Path

router = APIRouter(prefix="/reportes", tags=["Reportes"])


def obtener_datos_reporte(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    casino: Optional[str] = None,
    marca: Optional[str] = None,
    modelo: Optional[str] = None,
    maquinas: Optional[List[str]] = None,
):
    registros = []
    fecha_inicio_str = fecha_inicio.isoformat() if fecha_inicio else None
    fecha_fin_str = fecha_fin.isoformat() if fecha_fin else None

    if modelo and modelo.lower() in ["todos", "todas"]:
        modelo = None
    if marca and marca.lower() in ["todos", "todas"]:
        marca = None
    if maquinas and (any(m.lower() in ["todos", "todas"] for m in maquinas)):
        maquinas = None

    if modelo or marca:
        from pathlib import Path
        import json
        ruta_maquinas = Path(__file__).parent.parent / \
            "gestion_maquinas" / "maquinas.json"
        codigos_filtrados = []
        if ruta_maquinas.exists():
            with open(ruta_maquinas, "r", encoding="utf-8") as f:
                maquinas_data = json.load(f)
                for m in maquinas_data.values():
                    cumple_modelo = not modelo or m.get(
                        "modelo", "").lower() == modelo.lower()
                    cumple_marca = not marca or m.get(
                        "marca", "").lower() == marca.lower()
                    cumple_casino = not casino or m.get("casino", "").lower() == (
                        casino or "").lower() or casino == "Todos"
                    if cumple_modelo and cumple_marca and cumple_casino:
                        codigos_filtrados.append(m["codigo"])
        if maquinas:
            maquinas = list(set(maquinas) & set(codigos_filtrados))
        else:
            maquinas = codigos_filtrados

    if (casino is None or casino == "Todos") and (not maquinas or len(maquinas) == 0):
        lugares = cargar_lugares()
        for datos in lugares.values():
            casino_nombre = datos["nombre_casino"]
            try:
                req = CuadreCasinoRequest(
                    casino=casino_nombre, fecha_inicio=fecha_inicio_str, fecha_fin=fecha_fin_str)
                resultado = cuadre_casino(req)
                for m in resultado["detalle_maquinas"]:
                    registros.append({
                        "fecha_inicio": m["fecha_inicio"],
                        "fecha_fin": m["fecha_fin"],
                        "casino": casino_nombre,
                        "maquina": m["maquina"],
                        "in": m["total_in"],
                        "out": m["total_out"],
                        "jackpot": m["total_jackpot"],
                        "billetero": m["total_billetero"],
                        "utilidad": m["utilidad"],
                        "denominacion": m["denominacion"],
                        "contador_inicial": m["contador_inicial"],
                        "contador_final": m["contador_final"],
                    })
            except Exception as e:
                print(
                    f"Error al obtener cuadre del casino '{casino_nombre}': {e}")
                continue
    elif casino and (not maquinas or len(maquinas) == 0):
        req = CuadreCasinoRequest(
            casino=casino, fecha_inicio=fecha_inicio_str, fecha_fin=fecha_fin_str)
        resultado = cuadre_casino(req)
        for m in resultado["detalle_maquinas"]:
            registros.append({
                "fecha_inicio": m["fecha_inicio"],
                "fecha_fin": m["fecha_fin"],
                "casino": casino,
                "maquina": m["maquina"],
                "in": m["total_in"],
                "out": m["total_out"],
                "jackpot": m["total_jackpot"],
                "billetero": m["total_billetero"],
                "utilidad": m["utilidad"],
                "denominacion": m["denominacion"],
                "contador_inicial": m["contador_inicial"],
                "contador_final": m["contador_final"],
            })
    elif maquinas:
        for maquina in maquinas:
            req = BalanceRequest(
                fecha_inicio=fecha_inicio_str,
                fecha_fin=fecha_fin_str,
                casino=casino or "",
                maquina=maquina,
                id=maquina,
                denominacion=1.0
            )
            resultado = calcular_balance(req)
            for r in resultado["cuadres"]:
                registros.append({
                    "fecha_inicio": r["fecha_inicio"],
                    "fecha_fin": r["fecha_fin"],
                    "casino": r["casino"],
                    "maquina": r["maquina"],
                    "in": r["total_in"],
                    "out": r["total_out"],
                    "jackpot": r["total_jackpot"],
                    "billetero": r["total_billetero"],
                    "utilidad": r["utilidad"],
                    "contador_inicial": r["contador_inicial"],
                    "contador_final": r["contador_final"],
                })
    return registros


@router.get("/generar-reporte")
def generar_reporte(
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    casino: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    modelo: Optional[str] = Query(None),
    maquinas: Optional[List[str]] = Query(None),
):
    try:
        print(
            f"Parametros recibidos: fecha_inicio={fecha_inicio}, fecha_fin={fecha_fin}, casino={casino}, marca={marca}, modelo={modelo}, maquinas={maquinas}")
        registros = obtener_datos_reporte(
            fecha_inicio, fecha_fin, casino, marca, modelo, maquinas)
        print(f"Registros obtenidos: {registros}")
        return {"registros": registros}
    except Exception as e:
        print(f"Error en generar_reporte: {e}")
        return JSONResponse(status_code=500, content={"error": f"Error interno del servidor: {str(e)}"})


@router.get("/exportar-reporte")
def exportar_reporte(
    formato: str = Query("pdf"),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    casino: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    modelo: Optional[str] = Query(None),
    maquinas: Optional[List[str]] = Query(None),
):
    registros = obtener_datos_reporte(
        fecha_inicio, fecha_fin, casino, marca, modelo, maquinas)
    if not registros:
        return JSONResponse(status_code=404, content={"error": "No hay datos para exportar"})
    export_dir = os.path.join(os.path.dirname(__file__), "../../../exports")
    os.makedirs(export_dir, exist_ok=True)
    unique_id = uuid.uuid4().hex[:8]

    # Cargar configuración visual
    config_path = Path(__file__).parent.parent / "usuarios_configuracion" / "configuracion.json"
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = {}
    nombre_empresa = config.get("nombre_empresa", "Casino")
    color_primario = config.get("color_primario", "#1d4ed8")
    color_fondo = config.get("color_fondo", "#ffffff")
    logo_url = config.get("logo_url", None)
    # Convertir color primario a RGB para openpyxl/fpdf
    def hex_to_rgb(hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    rgb_primario = hex_to_rgb(color_primario)

    if formato == "excel" or formato == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        df = pd.DataFrame(registros).drop(
            columns=["contador_inicial", "contador_final"], errors="ignore")
        nombre_archivo = f"reporte_{unique_id}.xlsx"
        ruta_archivo = os.path.join(export_dir, nombre_archivo)
        df.to_excel(ruta_archivo, index=False, sheet_name="Reporte")
        wb = openpyxl.load_workbook(ruta_archivo)
        ws = wb.active
        # Encabezado con color y fuente
        fill = PatternFill(start_color=color_primario.lstrip('#'), end_color=color_primario.lstrip('#'), fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12)
        align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border
        # Ajustar ancho de columnas automáticamente al texto más largo de cada columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        # Título arriba
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{nombre_empresa} - Reporte de Contadores"
        title_cell.font = Font(bold=True, size=14, color=color_primario.lstrip('#'))
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(ruta_archivo)
    elif formato == "pdf":
        from PIL import Image
        nombre_archivo = f"reporte_{unique_id}.pdf"
        ruta_archivo = os.path.join(export_dir, nombre_archivo)
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        # Marca de agua: logo grande y centrado, con opacidad baja
        if logo_url:
            logo_path = Path(__file__).parent.parent / "usuarios_configuracion" / "media" / "logo.png"
            if logo_path.exists():
                # Procesar logo para opacidad baja
                img = Image.open(str(logo_path)).convert("RGBA")
                alpha = img.split()[3]
                alpha = alpha.point(lambda p: int(p * 0.15))  # 15% opacidad
                img.putalpha(alpha)
                temp_logo_path = os.path.join(export_dir, f"temp_logo_{unique_id}.png")
                img.save(temp_logo_path)
                # Calcular tamaño para centrar y cubrir gran parte de la página
                page_w, page_h = pdf.w - 2 * pdf.l_margin, pdf.h - 2 * pdf.t_margin
                logo_w = page_w * 0.7
                logo_h = page_h * 0.7
                logo_x = (pdf.w - logo_w) / 2
                logo_y = (pdf.h - logo_h) / 2
                pdf.image(temp_logo_path, x=logo_x, y=logo_y, w=logo_w, h=logo_h)
        # Encabezado y datos
        pdf.set_font("Arial", "B", 16)
        pdf.set_text_color(*rgb_primario)
        pdf.cell(0, 15, nombre_empresa, ln=True, align="C")
        pdf.ln(2)
        pdf.set_font("Arial", "B", 12)
        pdf.set_text_color(*rgb_primario)
        pdf.cell(0, 10, "Reporte de Contadores", ln=True, align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", size=9)
        headers = ["fecha_inicio", "fecha_fin", "casino", "maquina",
                   "in", "out", "jackpot", "billetero", "utilidad"]
        # Etiquetas de encabezado sin salto de línea, palabras juntas
        header_labels = [
            "Fecha inicio", "Fecha fin", "Casino", "Máquina",
            "In", "Out", "Jackpot", "Billetero", "Utilidad"
        ]
        # Calcular ancho de columnas considerando el texto más largo entre encabezado y datos
        col_widths = []
        for idx, h in enumerate(headers):
            max_header_width = pdf.get_string_width(header_labels[idx])
            max_val = max([pdf.get_string_width(str(r.get(h, ""))) for r in registros] + [max_header_width])
            col_widths.append(max(max_val + 4, 20))
        # Encabezados con color y sin saltos de línea
        pdf.set_fill_color(*rgb_primario)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 10)
        # Calcular el ancho total de la tabla
        total_table_width = sum(col_widths)
        # Calcular la posición X para centrar la tabla
        page_width = pdf.w - 2 * pdf.l_margin
        table_x = pdf.l_margin + (page_width - total_table_width) / 2
        pdf.set_x(table_x)
        for i, label in enumerate(header_labels):
            pdf.cell(col_widths[i], 8, label, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_font("Arial", size=9)
        pdf.set_text_color(0, 0, 0)
        for r in registros:
            pdf.set_x(table_x)
            for i, h in enumerate(headers):
                val = str(r.get(h, ""))
                # Si el texto es más largo que la celda, usar multi_cell
                if pdf.get_string_width(val) > col_widths[i] - 2:
                    x = pdf.get_x()
                    y = pdf.get_y()
                    pdf.multi_cell(col_widths[i], 8, val, border=1, align="C")
                    pdf.set_xy(x + col_widths[i], y)
                else:
                    pdf.cell(col_widths[i], 8, val, border=1, align="C")
            pdf.ln()
        pdf.output(ruta_archivo)
        # Eliminar logo temporal si se creó
        if logo_url and logo_path.exists():
            try:
                os.remove(temp_logo_path)
            except Exception:
                pass
    else:
        return JSONResponse(status_code=400, content={"error": "Formato no soportado"})
    return FileResponse(ruta_archivo, filename=nombre_archivo)


@router.get("/reporte-participacion")
def reporte_participacion(
    porcentaje: float = Query(..., gt=0, lt=100),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    casino: Optional[str] = Query(None),
    maquinas: Optional[List[str]] = Query(None),
):
    registros = obtener_datos_reporte(
        fecha_inicio, fecha_fin, casino, None, None, maquinas)
    utilidad_total = sum(r["utilidad"] for r in registros)
    valor_participacion = utilidad_total * (porcentaje / 100)
    return {
        "utilidad_total": utilidad_total,
        "valor_participacion": valor_participacion,
        "registros": registros,
    }
